from __future__ import annotations

import csv
from pathlib import Path

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from rag_pipeline.ingestion.schema import ExtractedContent


def extract_text_from_txt(path: Path) -> ExtractedContent:
    return ExtractedContent(text=path.read_text(encoding="utf-8", errors="ignore").strip(), metadata={"kind": "plain_text"})


def extract_text_from_md(path: Path) -> ExtractedContent:
    return ExtractedContent(text=path.read_text(encoding="utf-8", errors="ignore").strip(), metadata={"kind": "markdown"})


def extract_text_from_pdf(path: Path) -> ExtractedContent:
    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    text = "\n\n".join(page for page in pages if page).strip()
    return ExtractedContent(text=text, metadata={"kind": "pdf", "page_count": len(reader.pages)})


def extract_text_from_docx(path: Path) -> ExtractedContent:
    doc = DocxDocument(str(path))
    paragraphs = [paragraph.text.strip() for paragraph in doc.paragraphs if paragraph.text.strip()]
    text = "\n\n".join(paragraphs).strip()
    return ExtractedContent(text=text, metadata={"kind": "docx", "paragraph_count": len(paragraphs)})


def extract_text_from_html(path: Path) -> ExtractedContent:
    html = path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    title = soup.title.get_text(strip=True) if soup.title and soup.title.get_text(strip=True) else ""
    text = soup.get_text(separator="\n\n", strip=True)
    return ExtractedContent(text=text.strip(), metadata={"kind": "html", "title": title})


def extract_text_from_csv(path: Path) -> ExtractedContent:
    rows: list[str] = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            cleaned = [cell.strip() for cell in row if cell.strip()]
            if cleaned:
                rows.append(" | ".join(cleaned))
    return ExtractedContent(text="\n".join(rows).strip(), metadata={"kind": "csv", "row_count": len(rows)})


def extract_text_from_xlsx(path: Path) -> ExtractedContent:
    workbook = load_workbook(path, data_only=True, read_only=True)
    output: list[str] = []
    sheet_row_counts: dict[str, int] = {}
    for sheet in workbook.worksheets:
        output.append(f"[Sheet: {sheet.title}]")
        rows_seen = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                output.append(" | ".join(cells))
                rows_seen += 1
        sheet_row_counts[sheet.title] = rows_seen
    return ExtractedContent(text="\n".join(output).strip(), metadata={"kind": "xlsx", "sheet_row_counts": sheet_row_counts})


def extract_text_from_code(path: Path) -> ExtractedContent:
    return ExtractedContent(text=path.read_text(encoding="utf-8", errors="ignore").strip(), metadata={"kind": "code", "language": path.suffix.lower().lstrip(".")})
